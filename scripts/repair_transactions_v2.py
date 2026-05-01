"""
repair_transactions_v2.py
Reads exact per-player per-week deduction amounts from Excel Weekly Tracker,
deletes the old (wrong) match_deduction transactions, and re-inserts correct ones.
Also re-inserts expense_deduction transactions.
"""
import openpyxl, sys, json, difflib, requests, datetime
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8')

BASE = 'https://tzcernzuwtwgrsattjaw.supabase.co/rest/v1'
KEY  = 'sb_secret_nyfXWpX2nEWP8azMu-Ip5A_iON9D5if'
H    = {
    'apikey': KEY, 'Authorization': f'Bearer {KEY}',
    'Content-Type': 'application/json', 'Prefer': 'resolution=merge-duplicates',
}
H_DEL = {
    'apikey': KEY, 'Authorization': f'Bearer {KEY}',
    'Content-Type': 'application/json',
}

def sb_get(path):
    r = requests.get(f'{BASE}{path}', headers=H)
    r.raise_for_status()
    return r.json()

# ── Fetch Supabase data ──────────────────────────────────────────────────────
players    = sb_get('/players?select=id,display_name,type,status&limit=200')
weeks      = sb_get('/weeks?select=week_id,match_date,status,match_fee,label&status=eq.completed')
attendance = sb_get('/attendance?select=player_id,week_id,status,sponsor_player_id&limit=2000')

player_map   = {p['id']: p for p in players}
week_map     = {w['week_id']: w for w in weeks}

# Build date -> week_id map for matching Excel week dates
date_to_wid = {}
for w in weeks:
    date_to_wid[w['match_date']] = w['week_id']

print(f'Completed weeks:     {len(weeks)}')
print(f'Attendance records:  {len(attendance)}')
print()

# ── Step 1: Delete old migration match_deduction transactions ─────────────────
print('Deleting old match_deduction transactions (recorded_by=migration)...')
r = requests.delete(
    f'{BASE}/transactions?type=eq.match_deduction&recorded_by=eq.migration',
    headers=H_DEL,
)
if not r.ok:
    print(f'  ERROR: {r.status_code} {r.text[:400]}')
else:
    print('  Deleted OK')

# ── Step 2: Read Weekly Tracker from Excel ────────────────────────────────────
print()
print('Reading Excel Weekly Tracker...')
wb = openpyxl.load_workbook('Cricket_Corpus_Tracker_v9.xlsx', data_only=True)

# Try to find the Weekly Tracker sheet
sheet_names = wb.sheetnames
print(f'Sheets available: {sheet_names}')

# Find the weekly tracker sheet
ws_weekly = None
for name in sheet_names:
    if 'weekly' in name.lower() or 'tracker' in name.lower() or 'week' in name.lower():
        ws_weekly = wb[name]
        print(f'Using sheet: {name}')
        break

if not ws_weekly:
    # Try first sheet or the one that looks right
    for name in sheet_names:
        ws_weekly = wb[name]
        print(f'Falling back to sheet: {name}')
        break

rows = list(ws_weekly.iter_rows(values_only=True))

# Print first 8 rows to understand structure
print('\nFirst 8 rows of Weekly Tracker:')
for i, row in enumerate(rows[:8]):
    print(f'  Row {i+1}: {row[:20]}')

# ── Parse week headers from row 4 (index 3) ──────────────────────────────────
# Structure: Row 4 has week labels like 'Wk1  22-Feb-2026' every 2 columns starting at col 4
# Row 5 has column headers: #, Category, Player Name, Type, Attend(P/A), Deduction(₹), ...
# Row 6+: player data

# Find the header row with week info
header_row_idx = None
week_col_start = None
wk_headers = {}  # col_index -> (week_label, date_str, week_id)

for ridx, row in enumerate(rows[:10]):
    for cidx, cell in enumerate(row):
        if cell and isinstance(cell, str) and 'Wk' in str(cell):
            header_row_idx = ridx
            break
    if header_row_idx is not None:
        break

if header_row_idx is None:
    print('ERROR: Could not find week header row!')
    sys.exit(1)

print(f'\nWeek header row: {header_row_idx + 1}')
header_row = rows[header_row_idx]

# Parse week headers - each week spans 2 columns (Attend, Deduction)
# Find where weeks start
for cidx, cell in enumerate(header_row):
    if cell and isinstance(cell, str) and 'Wk' in str(cell):
        week_col_start = cidx
        break

print(f'Week columns start at column index: {week_col_start}')

# Parse each week header
# 'Wk1  22-Feb-2026' -> date 2026-02-22
import re
for cidx in range(week_col_start, len(header_row), 2):
    cell = header_row[cidx]
    if not cell:
        continue
    cell_str = str(cell).strip()
    if 'Wk' not in cell_str:
        break

    # Extract date from the label
    # Format: 'Wk1  22-Feb-2026' or 'Wk2  01-Mar-2026'
    date_match = re.search(r'(\d{1,2})-(\w{3})-(\d{4})', cell_str)
    if date_match:
        day, mon, year = date_match.groups()
        try:
            dt = datetime.datetime.strptime(f'{day}-{mon}-{year}', '%d-%b-%Y')
            date_str = dt.strftime('%Y-%m-%d')
            wid = date_to_wid.get(date_str)
            wk_headers[cidx] = {
                'label': cell_str.strip(),
                'date': date_str,
                'week_id': wid,
                'attend_col': cidx,
                'deduct_col': cidx + 1,
            }
            print(f'  Col {cidx}: {cell_str} -> {date_str} -> {wid}')
        except ValueError as e:
            print(f'  Could not parse date from: {cell_str!r} ({e})')

print(f'\nParsed {len(wk_headers)} week columns')

# ── Find data rows (player rows) ──────────────────────────────────────────────
# Row 5 (index 4) should be column headers, row 6+ are player data
# Player rows have: col0=row#, col2=player_name, col3=type

# Find the name column and type column
col_row_headers = rows[header_row_idx + 1]  # row after week headers
print(f'\nColumn headers row: {col_row_headers[:10]}')

# Identify player name column (usually col 2 or similar)
name_col = None
type_col = None
for cidx, cell in enumerate(col_row_headers):
    if cell and 'name' in str(cell).lower():
        name_col = cidx
    if cell and 'type' in str(cell).lower():
        type_col = cidx

print(f'Name column: {name_col}, Type column: {type_col}')

if name_col is None:
    name_col = 2  # default based on known structure
if type_col is None:
    type_col = 3

# ── Build fuzzy name -> player_id map ────────────────────────────────────────
def fuzzy_match(name, player_list, threshold=0.55):
    best_score, best_id = 0, None
    for p in player_list:
        score = difflib.SequenceMatcher(None, name.lower(), p['display_name'].lower()).ratio()
        if score > best_score:
            best_score, best_id = score, p['id']
    if best_score >= threshold:
        return best_id, best_score
    return None, best_score

# ── Parse player rows and collect match deductions ────────────────────────────
data_start_row = header_row_idx + 2  # skip header row + column label row
print(f'\nReading player data from row {data_start_row + 1} onwards...')

match_txns = []
unmatched_players = set()

for ridx in range(data_start_row, len(rows)):
    row = rows[ridx]

    # Skip empty rows
    if not row[name_col]:
        continue
    name = str(row[name_col]).strip()
    if not name or name.lower() in ('player name', 'total', 'subtotal'):
        continue

    # Skip non-corpus/new players (guests, PPM don't have deductions tracked here)
    player_type_cell = str(row[type_col]).strip() if row[type_col] else ''
    if 'corpus' not in player_type_cell.lower() and 'new' not in player_type_cell.lower():
        # Still try to match, corpus is in description like '💰 Corpus'
        pass

    # Fuzzy match player name
    pid, score = fuzzy_match(name, players)
    if not pid:
        if name not in unmatched_players:
            print(f'  UNMATCHED player: {name!r} (score={score:.2f})')
            unmatched_players.add(name)
        continue

    player = player_map[pid]
    if player['type'] not in ('corpus', 'new'):
        continue  # Only corpus/new pay match deductions

    # For each week, read deduction amount
    for cidx, wk_info in wk_headers.items():
        attend_col = wk_info['attend_col']
        deduct_col = wk_info['deduct_col']
        wid = wk_info['week_id']

        if wid is None:
            continue  # Week not in Supabase (not completed or not found)

        if deduct_col >= len(row):
            continue

        deduction = row[deduct_col]
        if deduction is None or deduction == '' or deduction == 0:
            continue

        try:
            deduction = float(deduction)
        except (TypeError, ValueError):
            continue

        if deduction <= 0:
            continue

        txn_id = f'TXN_DEDUCT_{wid}_{pid}'
        week = week_map[wid]

        match_txns.append({
            'id':            txn_id,
            'player_id':     pid,
            'tournament_id': 'TRN_001',
            'type':          'match_deduction',
            'direction':     'debit',
            'amount':        round(deduction, 2),
            'date':          wk_info['date'],
            'week_id':       wid,
            'description':   f"Match fee - {week['label']}",
            'recorded_by':   'migration',
            'receipt_ref':   '',
        })

print(f'\nMatch deduction txns from Excel: {len(match_txns)}')

# Dedup (same player+week can't have two rows in Excel, but guard it)
final_map = {}
for t in match_txns:
    tid = t['id']
    if tid in final_map:
        final_map[tid]['amount'] = round(final_map[tid]['amount'] + t['amount'], 2)
    else:
        final_map[tid] = dict(t)
all_new = list(final_map.values())
print(f'After dedup: {len(all_new)} transactions')

# ── Step 3: Also rebuild expense deductions ───────────────────────────────────
print()
print('Deleting old expense_deduction transactions (recorded_by=migration)...')
r = requests.delete(
    f'{BASE}/transactions?type=eq.expense_deduction&recorded_by=eq.migration',
    headers=H_DEL,
)
if not r.ok:
    print(f'  ERROR: {r.status_code} {r.text[:400]}')
else:
    print('  Deleted OK')

ws_exp = wb['Expense Deductions']
exp_rows = list(ws_exp.iter_rows(values_only=True))

player_expense = {}
for row in exp_rows[5:]:
    name = row[0]
    if not name or not isinstance(name, str):
        continue
    type_cell = str(row[1]) if row[1] else ''
    if 'corpus' not in type_cell.lower() and 'new' not in type_cell.lower():
        continue
    total = sum(float(v) for v in row[2:] if isinstance(v, (int, float)) and v)
    if total > 0:
        player_expense[name.strip()] = round(total, 4)

name_to_id = {}
for excel_name in player_expense:
    pid, score = fuzzy_match(excel_name, players)
    if pid:
        name_to_id[excel_name] = pid
    else:
        print(f'  UNMATCHED expense player: {excel_name!r} (score={score:.2f})')

expense_txns = []
for excel_name, exp_total in player_expense.items():
    pid = name_to_id.get(excel_name)
    if not pid:
        continue
    txn_id = f'TXN_EXP_DEDUCT_{pid}'
    expense_txns.append({
        'id':            txn_id,
        'player_id':     pid,
        'tournament_id': 'TRN_001',
        'type':          'expense_deduction',
        'direction':     'debit',
        'amount':        exp_total,
        'date':          '2026-03-08',
        'week_id':       'W_2026_03_08',
        'description':   'Equipment expense share (bat + balls)',
        'recorded_by':   'migration',
        'receipt_ref':   '',
    })

print(f'Expense deduction txns: {len(expense_txns)}')

# ── Step 4: Insert all new transactions ─────────────────────────────────────
all_insert = all_new + expense_txns
print(f'\nTotal transactions to insert: {len(all_insert)}')
print()

if not all_insert:
    print('Nothing to insert.')
else:
    chunk = 500
    for i in range(0, len(all_insert), chunk):
        batch = all_insert[i:i+chunk]
        r = requests.post(f'{BASE}/transactions', headers=H, json=batch)
        if not r.ok:
            print(f'ERROR inserting rows {i+1}-{i+len(batch)}: {r.status_code} {r.text[:400]}')
        else:
            print(f'Inserted rows {i+1}-{i+len(batch)} OK')

# ── Step 5: Verify against Excel ─────────────────────────────────────────────
print()
print('=== BALANCE VERIFICATION ===')

updated = sb_get('/player_balances?select=id,display_name,corpus_balance,total_paid,total_deducted&limit=200')
sb_bal  = {p['display_name'].lower(): p for p in updated}

ws2 = wb['Corpus Summary']
rows2 = list(ws2.iter_rows(values_only=True))

ok, mismatch, unmatched = 0, 0, 0
print(f'{"Player":<22} {"Excel":>10} {"Supabase":>10} {"Diff":>8}')
print('-' * 57)
for row in rows2[5:]:
    if not row[0] or not isinstance(row[0], (int, float)):
        continue
    name   = str(row[2]).strip() if row[2] else ''
    ex_bal = row[7]
    if not name or ex_bal is None:
        continue
    try:
        ex_bal = float(ex_bal)
    except (TypeError, ValueError):
        continue

    match = sb_bal.get(name.lower())
    if not match:
        for k, v in sb_bal.items():
            if name.lower() in k or k in name.lower():
                match = v
                break

    if not match:
        print(f'  {name:<22} {ex_bal:>10.2f}  {"NOT FOUND":>10}')
        unmatched += 1
        continue

    sb_b = float(match['corpus_balance'])
    diff = abs(sb_b - ex_bal)
    flag = '  <-- MISMATCH' if diff > 5 else ''
    print(f'  {name:<22} {ex_bal:>10.2f}  {sb_b:>10.2f}  {diff:>7.2f}{flag}')
    if diff > 5:
        mismatch += 1
    else:
        ok += 1

print()
print(f'OK: {ok}  |  Mismatch (>5): {mismatch}  |  Unmatched: {unmatched}')
