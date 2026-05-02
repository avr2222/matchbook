"""
repair_transactions_v4.py  — Final correct repair
Fixes:
  1. Expense double-count bug: use pre-computed total col (col 22) not sum(row[2:])
  2. Ashok(Raaga) mismatch: override fuzzy match to PLY_033 (Ashok adapala) not PLY_016
  3. Verification name overrides: Karthik Reddy = PLY_009 (Kranthi), Ashok(Raaga) = PLY_033
"""
import openpyxl, sys, difflib, requests, datetime, re, os
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8')

BASE = os.environ['SUPABASE_URL'].rstrip('/') + '/rest/v1'
KEY  = os.environ['SUPABASE_SERVICE_KEY']
H    = {
    'apikey': KEY, 'Authorization': f'Bearer {KEY}',
    'Content-Type': 'application/json', 'Prefer': 'resolution=merge-duplicates',
}
H_DEL = {
    'apikey': KEY, 'Authorization': f'Bearer {KEY}',
    'Content-Type': 'application/json',
}

def sb_get(path):
    r = requests.get(f'{BASE}{path}', headers=H)
    r.raise_for_status()
    return r.json()

# ── Fetch Supabase data ──────────────────────────────────────────────────────
players    = sb_get('/players?select=id,display_name,type,status&limit=200')
weeks      = sb_get('/weeks?select=week_id,match_date,status,match_fee,label&status=eq.completed')
attendance = sb_get('/attendance?select=player_id,week_id,status,sponsor_player_id&limit=2000')

player_map   = {p['id']: p for p in players}
week_map     = {w['week_id']: w for w in weeks}
date_to_wid  = {w['match_date']: w['week_id'] for w in weeks}

print(f'Completed weeks: {len(weeks)}, Players: {len(players)}, Attendance: {len(attendance)}')

# Manual overrides: Excel display_name -> Supabase player_id
# 'Ashok(Raaga)' in Excel = PLY_033 (display_name='Ashok adapala') in Supabase
EXCEL_TO_PID_OVERRIDES = {
    'Ashok(Raaga)': 'PLY_033',
}

# ── Step 1: Delete old wrong transactions ────────────────────────────────────
print('\nDeleting old match_deduction transactions (recorded_by=migration)...')
r = requests.delete(f'{BASE}/transactions?type=eq.match_deduction&recorded_by=eq.migration', headers=H_DEL)
print(f'  {r.status_code}' + ('' if r.ok else f' {r.text[:200]}'))

print('Deleting old expense_deduction transactions (recorded_by=migration)...')
r = requests.delete(f'{BASE}/transactions?type=eq.expense_deduction&recorded_by=eq.migration', headers=H_DEL)
print(f'  {r.status_code}' + ('' if r.ok else f' {r.text[:200]}'))

# ── Step 2: Read Excel Weekly Tracker ────────────────────────────────────────
wb = openpyxl.load_workbook('Cricket_Corpus_Tracker_v9.xlsx', data_only=True)
ws = wb['Weekly Tracker']
rows = list(ws.iter_rows(values_only=True))

HEADER_ROW     = 3   # 0-indexed (Row 4 in Excel)
DATA_START_ROW = 5   # 0-indexed (Row 6 in Excel)
NAME_COL       = 2
TYPE_COL       = 3
WEEK_START_COL = 4

hrow = rows[HEADER_ROW]

def fuzzy_match(name, player_list, threshold=0.55):
    best_score, best_id = 0, None
    for p in player_list:
        score = difflib.SequenceMatcher(None, name.lower(), p['display_name'].lower()).ratio()
        if score > best_score:
            best_score, best_id = score, p['id']
    return (best_id, best_score) if best_score >= threshold else (None, best_score)

def resolve_excel_name(name):
    """Return Supabase player_id for an Excel player name, using overrides first."""
    if name in EXCEL_TO_PID_OVERRIDES:
        return EXCEL_TO_PID_OVERRIDES[name]
    pid, _ = fuzzy_match(name, players)
    return pid

# Parse week columns from row 4
excel_weeks = {}
for cidx in range(WEEK_START_COL, len(hrow), 2):
    cell = hrow[cidx]
    if not cell or 'Wk' not in str(cell):
        break
    m = re.search(r'(\d{1,2})-(\w{3})-(\d{4})', str(cell))
    if not m:
        continue
    day, mon, yr = m.groups()
    try:
        dt = datetime.datetime.strptime(f'{day}-{mon}-{yr}', '%d-%b-%Y')
        date_str = dt.strftime('%Y-%m-%d')
        wid = date_to_wid.get(date_str)
        excel_weeks[cidx] = {
            'label': str(cell).strip(), 'date': date_str,
            'week_id': wid, 'attend_col': cidx, 'deduct_col': cidx + 1,
        }
    except ValueError:
        pass

matched_wks = sum(1 for w in excel_weeks.values() if w['week_id'])
print(f'\nExcel weeks: {len(excel_weeks)} found, {matched_wks} matched to Supabase completed weeks')

# ── Step 3: Read per-player per-week deductions from Excel ───────────────────
print('Reading Weekly Tracker deductions...')
match_txns_excel = []
seen = set()

for ridx in range(DATA_START_ROW, len(rows)):
    row = rows[ridx]
    name_cell = row[NAME_COL]
    if not name_cell or not isinstance(name_cell, str):
        continue
    name = name_cell.strip()
    if not name or name.lower() in ('player name', '[player name]', 'total', 'subtotal'):
        continue

    pid = resolve_excel_name(name)
    if not pid:
        continue

    player = player_map.get(pid)
    if not player or player['type'] not in ('corpus', 'new'):
        continue

    for cidx, wk_info in excel_weeks.items():
        wid = wk_info['week_id']
        if not wid:
            continue
        deduct_col = wk_info['deduct_col']
        if deduct_col >= len(row):
            continue
        try:
            deduction = float(row[deduct_col])
        except (TypeError, ValueError):
            continue
        if deduction <= 0:
            continue

        week = week_map[wid]
        key = (wid, pid)
        if key in seen:
            # Merge amounts for same player+week (shouldn't happen with overrides)
            for t in match_txns_excel:
                if t['week_id'] == wid and t['player_id'] == pid:
                    t['amount'] = round(t['amount'] + deduction, 2)
                    break
        else:
            seen.add(key)
            match_txns_excel.append({
                'id':            f'TXN_DEDUCT_{wid}_{pid}',
                'player_id':     pid,
                'tournament_id': 'TRN_001',
                'type':          'match_deduction',
                'direction':     'debit',
                'amount':        round(deduction, 2),
                'date':          wk_info['date'],
                'week_id':       wid,
                'description':   f"Match fee - {week['label']}",
                'recorded_by':   'migration',
                'receipt_ref':   '',
            })

print(f'  Excel match deduction rows: {len(match_txns_excel)}')

# ── Step 4: Supabase-only weeks (not yet in Excel) ────────────────────────────
excel_covered = {w['week_id'] for w in excel_weeks.values() if w['week_id']}
supabase_only = [w for w in weeks if w['week_id'] not in excel_covered]
print(f'\nSupabase-only weeks (no Excel data yet): {len(supabase_only)}')
for w in supabase_only:
    print(f'  {w["week_id"]} | {w["match_date"]} | fee={w["match_fee"]} | {w["label"]}')

match_txns_sb = []
for week in supabase_only:
    wid = week['week_id']
    charges = defaultdict(float)
    for rec in attendance:
        if rec['week_id'] != wid or rec['status'] != 'played':
            continue
        pid    = rec['player_id']
        player = player_map.get(pid)
        if not player:
            continue
        ptype = player['type']
        if ptype in ('corpus', 'new'):
            charges[pid] += float(week['match_fee'])
        elif ptype == 'guest' and rec.get('sponsor_player_id'):
            charges[rec['sponsor_player_id']] += float(week['match_fee'])

    for charge_pid, amount in charges.items():
        match_txns_sb.append({
            'id':            f'TXN_DEDUCT_{wid}_{charge_pid}',
            'player_id':     charge_pid,
            'tournament_id': 'TRN_001',
            'type':          'match_deduction',
            'direction':     'debit',
            'amount':        round(amount, 2),
            'date':          week['match_date'],
            'week_id':       wid,
            'description':   f"Match fee - {week['label']}",
            'recorded_by':   'migration',
            'receipt_ref':   '',
        })

print(f'  Supabase-sourced match deduction rows: {len(match_txns_sb)}')

# ── Step 5: Expense deductions (use col 22 = pre-computed total) ─────────────
print('\nReading expense deductions from Excel Expense Deductions sheet...')
EXPENSE_TOTAL_COL = 22  # Column W = pre-computed total, avoids double-count from sum(row[2:])

ws_exp = wb['Expense Deductions']
exp_rows = list(ws_exp.iter_rows(values_only=True))

expense_txns = []
for row in exp_rows[5:]:
    name = row[0]
    if not name or not isinstance(name, str):
        continue
    name = name.strip()
    type_cell = str(row[1]) if row[1] else ''
    if 'corpus' not in type_cell.lower() and 'new' not in type_cell.lower():
        continue

    # Use the pre-computed total column directly
    if len(row) <= EXPENSE_TOTAL_COL or row[EXPENSE_TOTAL_COL] is None:
        continue
    try:
        total = float(row[EXPENSE_TOTAL_COL])
    except (TypeError, ValueError):
        continue
    if total <= 0:
        continue

    pid = resolve_excel_name(name)
    if not pid:
        continue

    expense_txns.append({
        'id':            f'TXN_EXP_DEDUCT_{pid}',
        'player_id':     pid,
        'tournament_id': 'TRN_001',
        'type':          'expense_deduction',
        'direction':     'debit',
        'amount':        round(total, 4),
        'date':          '2026-03-08',
        'week_id':       'W_2026_03_08',
        'description':   'Equipment expense share (bat + balls)',
        'recorded_by':   'migration',
        'receipt_ref':   '',
    })

print(f'  Expense deduction rows: {len(expense_txns)}')

# ── Step 6: Dedup and insert ──────────────────────────────────────────────────
all_txns = match_txns_excel + match_txns_sb + expense_txns
final_map = {}
for t in all_txns:
    tid = t['id']
    if tid in final_map:
        final_map[tid]['amount'] = round(final_map[tid]['amount'] + t['amount'], 2)
    else:
        final_map[tid] = dict(t)

all_new = list(final_map.values())
print(f'\nTotal transactions to insert: {len(all_new)}')
print(f'  Match deductions (Excel):    {len(match_txns_excel)}')
print(f'  Match deductions (Supabase): {len(match_txns_sb)}')
print(f'  Expense deductions:          {len(expense_txns)}')
print()

if not all_new:
    print('Nothing to insert.')
else:
    chunk = 500
    for i in range(0, len(all_new), chunk):
        batch = all_new[i:i+chunk]
        r = requests.post(f'{BASE}/transactions', headers=H, json=batch)
        if not r.ok:
            print(f'ERROR {i+1}-{i+len(batch)}: {r.status_code} {r.text[:400]}')
        else:
            print(f'Inserted rows {i+1}-{i+len(batch)} OK')

# ── Step 7: Verify ───────────────────────────────────────────────────────────
print()
print('=== BALANCE VERIFICATION ===')
print('Note: players who played Apr 25 will show extra ₹500 vs Excel (Excel not yet updated for that match)')
print()

updated = sb_get('/player_balances?select=id,display_name,corpus_balance,total_paid,total_deducted&limit=200')
sb_bal_by_id   = {p['id']: p for p in updated}
sb_bal_by_name = {p['display_name'].lower(): p for p in updated}

# Corpus Summary name → Supabase player_id overrides
VERIFY_OVERRIDES = {
    'karthik reddy': 'PLY_009',   # Corpus Summary calls PLY_009 (Kranthi) as 'Karthik Reddy'
    'ashok(raaga)':  'PLY_033',   # Corpus Summary's Ashok(Raaga) = PLY_033 (Ashok adapala)
}

apr25_wid = 'W_2026_04_25'
apr25_players = set()
for rec in attendance:
    if rec['week_id'] == apr25_wid and rec['status'] == 'played':
        pid = rec['player_id']
        p = player_map.get(pid)
        if p and p['type'] in ('corpus', 'new'):
            apr25_players.add(pid)

apr25_fee = float(week_map.get(apr25_wid, {}).get('match_fee', 0) or 0)

ws2 = wb['Corpus Summary']
rows2 = list(ws2.iter_rows(values_only=True))

ok, mismatch, unmatched_v = 0, 0, 0
print(f'  {"Player":<22} {"ExcelBal":>10} {"SBBal":>10} {"Diff":>8}  Note')
print('  ' + '-' * 72)

for row in rows2[5:]:
    if not row[0] or not isinstance(row[0], (int, float)):
        continue
    name = str(row[2]).strip() if row[2] else ''
    ex_bal = row[7]
    if not name or ex_bal is None:
        continue
    try:
        ex_bal = float(ex_bal)
    except (TypeError, ValueError):
        continue

    # Resolve Supabase player
    pid = VERIFY_OVERRIDES.get(name.lower())
    if pid:
        sb_p = sb_bal_by_id.get(pid)
    else:
        sb_p = sb_bal_by_name.get(name.lower())
        if not sb_p:
            for k, v in sb_bal_by_name.items():
                if name.lower() in k or k in name.lower():
                    sb_p = v
                    break

    if not sb_p:
        print(f'  {name:<22} {ex_bal:>10.2f}  {"NOT FOUND":>10}')
        unmatched_v += 1
        continue

    pid_resolved = sb_p['id']
    sb_b = float(sb_p['corpus_balance'])
    diff = abs(sb_b - ex_bal)

    played_apr25 = pid_resolved in apr25_players
    note = f'(played Apr25, -₹{apr25_fee:.0f} vs Excel)' if played_apr25 else ''

    is_mismatch = diff > 5 and not played_apr25
    flag = '  <-- MISMATCH' if is_mismatch else ''
    print(f'  {name:<22} {ex_bal:>10.2f}  {sb_b:>10.2f}  {diff:>7.2f}{flag}  {note}')
    if is_mismatch:
        mismatch += 1
    else:
        ok += 1

print()
print(f'  OK: {ok}  |  Mismatch (>5 excluding Apr25): {mismatch}  |  Unmatched: {unmatched_v}')
