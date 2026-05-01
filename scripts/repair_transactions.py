import openpyxl, sys, json, difflib, requests
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8')

BASE = 'https://tzcernzuwtwgrsattjaw.supabase.co/rest/v1'
KEY  = 'sb_secret_nyfXWpX2nEWP8azMu-Ip5A_iON9D5if'
H    = {
    'apikey': KEY, 'Authorization': f'Bearer {KEY}',
    'Content-Type': 'application/json', 'Prefer': 'resolution=merge-duplicates',
}

def sb_get(path):
    r = requests.get(f'{BASE}{path}', headers=H)
    r.raise_for_status()
    return r.json()

# ── Fetch Supabase data ──────────────────────────────────────────────────────
players    = sb_get('/players?select=id,display_name,type,status&limit=200')
weeks      = sb_get('/weeks?select=week_id,match_date,status,match_fee,label&status=eq.completed')
attendance = sb_get('/attendance?select=player_id,week_id,status,sponsor_player_id&limit=1000')
existing   = sb_get('/transactions?select=id&limit=1000')

existing_ids = {t['id'] for t in existing}
player_map   = {p['id']: p for p in players}
week_map     = {w['week_id']: w for w in weeks}

print(f'Completed weeks:      {len(weeks)}')
print(f'Attendance records:   {len(attendance)}')
print(f'Existing transactions: {len(existing_ids)}')
print()

# ── Step 1: Match deduction transactions ────────────────────────────────────
# Accumulate per (week, charge_player) to handle multiple guests → same sponsor
charges = defaultdict(float)
charge_meta = {}

for rec in attendance:
    if rec['status'] != 'played':
        continue
    pid    = rec['player_id']
    wid    = rec['week_id']
    week   = week_map.get(wid)
    player = player_map.get(pid)
    if not week or not player:
        continue

    ptype = player['type']
    if ptype in ('corpus', 'new'):
        charge_pid = pid
    elif ptype == 'guest' and rec.get('sponsor_player_id'):
        charge_pid = rec['sponsor_player_id']
    else:
        continue  # PPM or unsponsored guest

    key = (wid, charge_pid)
    charges[key] += float(week['match_fee'])
    charge_meta[key] = week

match_txns = []
for (wid, charge_pid), amount in charges.items():
    week   = charge_meta[(wid, charge_pid)]
    txn_id = f'TXN_DEDUCT_{wid}_{charge_pid}'
    if txn_id in existing_ids:
        continue
    match_txns.append({
        'id':            txn_id,
        'player_id':     charge_pid,
        'tournament_id': 'TRN_001',
        'type':          'match_deduction',
        'direction':     'debit',
        'amount':        round(amount, 2),
        'date':          week['match_date'],
        'week_id':       wid,
        'description':   f"Match fee - {week['label']}",
        'recorded_by':   'migration',
        'receipt_ref':   '',
    })

print(f'Match deduction txns to insert: {len(match_txns)}')

# ── Step 2: Expense deductions from Excel ───────────────────────────────────
wb = openpyxl.load_workbook('Cricket_Corpus_Tracker_v9.xlsx', data_only=True)
ws = wb['Expense Deductions']
rows = list(ws.iter_rows(values_only=True))

player_expense = {}
for row in rows[5:]:
    name = row[0]
    if not name or not isinstance(name, str):
        continue
    type_cell = str(row[1]) if row[1] else ''
    if 'corpus' not in type_cell.lower() and 'new' not in type_cell.lower():
        continue
    total = sum(float(v) for v in row[2:] if isinstance(v, (int, float)) and v)
    if total > 0:
        player_expense[name.strip()] = round(total, 4)

# Fuzzy match Excel names → Supabase player IDs
name_to_id = {}
for excel_name in player_expense:
    best_score, best_id = 0, None
    for p in players:
        score = difflib.SequenceMatcher(None, excel_name.lower(), p['display_name'].lower()).ratio()
        if score > best_score:
            best_score, best_id = score, p['id']
    if best_score >= 0.55:
        name_to_id[excel_name] = best_id
    else:
        print(f'  UNMATCHED expense player: {excel_name!r} (best={best_score:.2f})')

expense_txns = []
for excel_name, exp_total in player_expense.items():
    pid    = name_to_id.get(excel_name)
    if not pid:
        continue
    txn_id = f'TXN_EXP_DEDUCT_{pid}'
    if txn_id in existing_ids:
        continue
    expense_txns.append({
        'id':            txn_id,
        'player_id':     pid,
        'tournament_id': 'TRN_001',
        'type':          'expense_deduction',
        'direction':     'debit',
        'amount':        exp_total,
        'date':          '2026-03-08',
        'week_id':       'W_2026_03_08',
        'description':   'Equipment expense share (bat + balls)',
        'recorded_by':   'migration',
        'receipt_ref':   '',
    })

print(f'Expense deduction txns to insert: {len(expense_txns)}')

# ── Step 3: Insert ───────────────────────────────────────────────────────────
# Final dedup: if two rows share an ID (shouldn't happen but guard it),
# keep the one with the highest amount (e.g. two guests → same sponsor same week)
final_map = {}
for t in match_txns + expense_txns:
    tid = t['id']
    if tid in final_map:
        final_map[tid]['amount'] = round(final_map[tid]['amount'] + t['amount'], 2)
    else:
        final_map[tid] = dict(t)
all_new = list(final_map.values())
print(f'Total new transactions:          {len(all_new)}')
print()

if not all_new:
    print('Nothing to insert.')
else:
    chunk = 500
    for i in range(0, len(all_new), chunk):
        batch = all_new[i:i+chunk]
        r = requests.post(f'{BASE}/transactions', headers=H, json=batch)
        if not r.ok:
            print(f'ERROR inserting rows {i+1}-{i+len(batch)}: {r.status_code} {r.text[:400]}')
        else:
            print(f'Inserted rows {i+1}-{i+len(batch)} OK')

# ── Step 4: Verify against Excel ────────────────────────────────────────────
print()
print('=== BALANCE VERIFICATION ===')

# Re-fetch updated balances
updated = sb_get('/player_balances?select=id,display_name,corpus_balance,total_paid,total_deducted&limit=200')
sb_bal  = {p['display_name'].lower(): p for p in updated}

# Read Corpus Summary for ground truth
ws2 = wb['Corpus Summary']
rows2 = list(ws2.iter_rows(values_only=True))

ok, mismatch, unmatched = 0, 0, 0
print(f'{"Player":<22} {"Excel":>10} {"Supabase":>10} {"Diff":>8}')
print('-' * 55)
for row in rows2[5:]:
    if not row[0] or not isinstance(row[0], (int, float)):
        continue
    name   = str(row[2]).strip() if row[2] else ''
    ex_bal = row[7]
    if not name or ex_bal is None:
        continue
    try:
        ex_bal = float(ex_bal)
    except (TypeError, ValueError):
        continue

    # Try to find in Supabase
    match = sb_bal.get(name.lower())
    if not match:
        # Try partial match
        for k, v in sb_bal.items():
            if name.lower() in k or k in name.lower():
                match = v
                break

    if not match:
        print(f'  {name:<22} {ex_bal:>10.2f}  {"NOT FOUND":>10}')
        unmatched += 1
        continue

    sb_b = float(match['corpus_balance'])
    diff = abs(sb_b - ex_bal)
    flag = '  <-- MISMATCH' if diff > 5 else ''
    print(f'  {name:<22} {ex_bal:>10.2f}  {sb_b:>10.2f}  {diff:>7.2f}{flag}')
    if diff > 5:
        mismatch += 1
    else:
        ok += 1

print()
print(f'OK: {ok}  |  Mismatch (>₹5): {mismatch}  |  Unmatched: {unmatched}')
