"""
repair_transactions_v3.py
- Deletes wrong match_deduction and expense_deduction transactions (recorded_by=migration)
- For weeks Wk1-Wk7: inserts exact per-player deductions read from Excel Weekly Tracker
- For Apr 25 week (not yet in Excel): inserts using Supabase attendance * match_fee
- Re-inserts expense_deductions from Excel Expense Deductions sheet
- Verifies final balances against Excel Corpus Summary
"""
import openpyxl, sys, difflib, requests, datetime, re
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8')

BASE = 'https://tzcernzuwtwgrsattjaw.supabase.co/rest/v1'
KEY  = 'sb_secret_nyfXWpX2nEWP8azMu-Ip5A_iON9D5if'
H    = {
    'apikey': KEY, 'Authorization': f'Bearer {KEY}',
    'Content-Type': 'application/json', 'Prefer': 'resolution=merge-duplicates',
}
H_DEL = {
    'apikey': KEY, 'Authorization': f'Bearer {KEY}',
    'Content-Type': 'application/json',
}

def sb_get(path):
    r = requests.get(f'{BASE}{path}', headers=H)
    r.raise_for_status()
    return r.json()

# ── Fetch Supabase data ──────────────────────────────────────────────────────
players    = sb_get('/players?select=id,display_name,type,status&limit=200')
weeks      = sb_get('/weeks?select=week_id,match_date,status,match_fee,label&status=eq.completed')
attendance = sb_get('/attendance?select=player_id,week_id,status,sponsor_player_id&limit=2000')

player_map   = {p['id']: p for p in players}
week_map     = {w['week_id']: w for w in weeks}
date_to_wid  = {w['match_date']: w['week_id'] for w in weeks}

print(f'Completed weeks:     {len(weeks)}')
print(f'Players:             {len(players)}')
print(f'Attendance records:  {len(attendance)}')
print()

# ── Step 1: Delete old wrong transactions ────────────────────────────────────
print('Deleting old match_deduction transactions (recorded_by=migration)...')
r = requests.delete(
    f'{BASE}/transactions?type=eq.match_deduction&recorded_by=eq.migration',
    headers=H_DEL
)
print(f'  Status: {r.status_code}' + ('' if r.ok else f' {r.text[:200]}'))

print('Deleting old expense_deduction transactions (recorded_by=migration)...')
r = requests.delete(
    f'{BASE}/transactions?type=eq.expense_deduction&recorded_by=eq.migration',
    headers=H_DEL
)
print(f'  Status: {r.status_code}' + ('' if r.ok else f' {r.text[:200]}'))
print()

# ── Step 2: Read Excel Weekly Tracker ────────────────────────────────────────
print('Reading Excel Weekly Tracker...')
wb = openpyxl.load_workbook('Cricket_Corpus_Tracker_v9.xlsx', data_only=True)
ws = wb['Weekly Tracker']
rows = list(ws.iter_rows(values_only=True))

# Row 4 (index 3): week headers at every other column starting at col 4
# Row 5 (index 4): column headers (#, Category, Player Name, Type, Attend/Deduct pairs)
# Row 6+ (index 5+): player data
HEADER_ROW = 3
DATA_START  = 5
NAME_COL    = 2
TYPE_COL    = 3
WEEK_START_COL = 4

hrow = rows[HEADER_ROW]

# Parse week headers: col 4='Wk1  22-Feb-2026', col 6='Wk2 01-Mar-2026', etc.
excel_weeks = {}  # attend_col -> {label, date, week_id}
for cidx in range(WEEK_START_COL, len(hrow), 2):
    cell = hrow[cidx]
    if not cell or 'Wk' not in str(cell):
        break
    m = re.search(r'(\d{1,2})-(\w{3})-(\d{4})', str(cell))
    if not m:
        continue
    day, mon, yr = m.groups()
    try:
        dt = datetime.datetime.strptime(f'{day}-{mon}-{yr}', '%d-%b-%Y')
        date_str = dt.strftime('%Y-%m-%d')
        wid = date_to_wid.get(date_str)
        excel_weeks[cidx] = {
            'label':      str(cell).strip(),
            'date':       date_str,
            'week_id':    wid,        # None if week not in Supabase
            'attend_col': cidx,
            'deduct_col': cidx + 1,
        }
    except ValueError:
        pass

print(f'Excel week columns found: {len(excel_weeks)}')
matched = sum(1 for w in excel_weeks.values() if w['week_id'])
print(f'  Matched to Supabase weeks: {matched}')
for cidx, wk in sorted(excel_weeks.items()):
    status = wk['week_id'] or 'NOT IN SUPABASE'
    print(f'  col {cidx}: {wk["label"]} -> {status}')
print()

# ── Fuzzy match helper ────────────────────────────────────────────────────────
def fuzzy_match(name, player_list, threshold=0.55):
    best_score, best_id = 0, None
    for p in player_list:
        score = difflib.SequenceMatcher(None, name.lower(), p['display_name'].lower()).ratio()
        if score > best_score:
            best_score, best_id = score, p['id']
    return (best_id, best_score) if best_score >= threshold else (None, best_score)

# ── Step 3: Parse player rows and collect Excel-based match deductions ────────
print('Reading per-player per-week deductions from Excel...')
match_txns_excel = []
unmatched = set()

for ridx in range(DATA_START, len(rows)):
    row = rows[ridx]
    name_cell = row[NAME_COL]
    if not name_cell or not isinstance(name_cell, str):
        continue
    name = name_cell.strip()
    if not name or name.lower() in ('player name', 'total', 'subtotal', '[player name]'):
        continue

    # Fuzzy match to Supabase player
    pid, score = fuzzy_match(name, players)
    if not pid:
        if name not in unmatched:
            print(f'  UNMATCHED: {name!r} (score={score:.2f})')
            unmatched.add(name)
        continue

    player = player_map[pid]
    if player['type'] not in ('corpus', 'new'):
        continue  # PPM players don't have corpus deductions

    # Read deduction for each week
    for cidx, wk_info in excel_weeks.items():
        wid = wk_info['week_id']
        if not wid:
            continue  # Week not in Supabase

        deduct_col = wk_info['deduct_col']
        if deduct_col >= len(row):
            continue

        deduction = row[deduct_col]
        if deduction is None or deduction == '' or deduction == 0:
            continue
        try:
            deduction = float(deduction)
        except (TypeError, ValueError):
            continue
        if deduction <= 0:
            continue

        week = week_map[wid]
        match_txns_excel.append({
            'id':            f'TXN_DEDUCT_{wid}_{pid}',
            'player_id':     pid,
            'tournament_id': 'TRN_001',
            'type':          'match_deduction',
            'direction':     'debit',
            'amount':        round(deduction, 2),
            'date':          wk_info['date'],
            'week_id':       wid,
            'description':   f"Match fee - {week['label']}",
            'recorded_by':   'migration',
            'receipt_ref':   '',
        })

print(f'Excel-sourced match deduction rows: {len(match_txns_excel)}')

# ── Step 4: Add Apr 25 deductions from Supabase (not yet in Excel) ────────────
# Find Supabase weeks NOT covered by Excel data
excel_covered_wids = {w['week_id'] for w in excel_weeks.values() if w['week_id']}
supabase_only_weeks = [w for w in weeks if w['week_id'] not in excel_covered_wids]
print(f'\nSupabase-only weeks (not in Excel yet): {len(supabase_only_weeks)}')
for w in supabase_only_weeks:
    print(f'  {w["week_id"]} | {w["match_date"]} | fee={w["match_fee"]} | {w["label"]}')

# For Supabase-only weeks, use attendance + match_fee
match_txns_supabase = []
for week in supabase_only_weeks:
    wid = week['week_id']
    # Accumulate per (week, charge_player_id) to handle guests with same sponsor
    charges = defaultdict(float)
    for rec in attendance:
        if rec['week_id'] != wid or rec['status'] != 'played':
            continue
        pid    = rec['player_id']
        player = player_map.get(pid)
        if not player:
            continue

        ptype = player['type']
        if ptype in ('corpus', 'new'):
            charge_pid = pid
        elif ptype == 'guest' and rec.get('sponsor_player_id'):
            charge_pid = rec['sponsor_player_id']
        else:
            continue

        charges[charge_pid] += float(week['match_fee'])

    for charge_pid, amount in charges.items():
        match_txns_supabase.append({
            'id':            f'TXN_DEDUCT_{wid}_{charge_pid}',
            'player_id':     charge_pid,
            'tournament_id': 'TRN_001',
            'type':          'match_deduction',
            'direction':     'debit',
            'amount':        round(amount, 2),
            'date':          week['match_date'],
            'week_id':       wid,
            'description':   f"Match fee - {week['label']}",
            'recorded_by':   'migration',
            'receipt_ref':   '',
        })

print(f'Supabase-sourced match deduction rows: {len(match_txns_supabase)}')

# ── Step 5: Rebuild expense deductions ───────────────────────────────────────
print()
print('Reading expense deductions from Excel...')
ws_exp = wb['Expense Deductions']
exp_rows = list(ws_exp.iter_rows(values_only=True))

player_expense = {}
for row in exp_rows[5:]:
    name = row[0]
    if not name or not isinstance(name, str):
        continue
    type_cell = str(row[1]) if row[1] else ''
    if 'corpus' not in type_cell.lower() and 'new' not in type_cell.lower():
        continue
    total = sum(float(v) for v in row[2:] if isinstance(v, (int, float)) and v)
    if total > 0:
        player_expense[name.strip()] = round(total, 4)

name_to_id = {}
for excel_name in player_expense:
    pid, score = fuzzy_match(excel_name, players)
    if pid:
        name_to_id[excel_name] = pid
    else:
        print(f'  UNMATCHED expense player: {excel_name!r} (score={score:.2f})')

expense_txns = []
for excel_name, exp_total in player_expense.items():
    pid = name_to_id.get(excel_name)
    if not pid:
        continue
    expense_txns.append({
        'id':            f'TXN_EXP_DEDUCT_{pid}',
        'player_id':     pid,
        'tournament_id': 'TRN_001',
        'type':          'expense_deduction',
        'direction':     'debit',
        'amount':        exp_total,
        'date':          '2026-03-08',
        'week_id':       'W_2026_03_08',
        'description':   'Equipment expense share (bat + balls)',
        'recorded_by':   'migration',
        'receipt_ref':   '',
    })

print(f'Expense deduction rows: {len(expense_txns)}')

# ── Step 6: Dedup and insert ──────────────────────────────────────────────────
all_txns = match_txns_excel + match_txns_supabase + expense_txns

# Dedup: same id from multiple sources -> merge amounts
final_map = {}
for t in all_txns:
    tid = t['id']
    if tid in final_map:
        final_map[tid]['amount'] = round(final_map[tid]['amount'] + t['amount'], 2)
    else:
        final_map[tid] = dict(t)

all_new = list(final_map.values())
print(f'\nTotal transactions to insert: {len(all_new)}')
print(f'  Match deductions (Excel):    {len(match_txns_excel)}')
print(f'  Match deductions (Supabase): {len(match_txns_supabase)}')
print(f'  Expense deductions:          {len(expense_txns)}')
print()

if not all_new:
    print('Nothing to insert.')
else:
    chunk = 500
    for i in range(0, len(all_new), chunk):
        batch = all_new[i:i+chunk]
        r = requests.post(f'{BASE}/transactions', headers=H, json=batch)
        if not r.ok:
            print(f'ERROR inserting rows {i+1}-{i+len(batch)}: {r.status_code} {r.text[:400]}')
        else:
            print(f'Inserted rows {i+1}-{i+len(batch)} OK')

# ── Step 7: Verify against Excel Corpus Summary ───────────────────────────────
print()
print('=== BALANCE VERIFICATION (Excel Corpus Summary vs Supabase) ===')
print('Note: Excel has data through Wk7 (05-Apr-2026).')
print('      Players who played Apr 25 will show extra deductions in Supabase.')
print()

updated = sb_get('/player_balances?select=id,display_name,corpus_balance,total_paid,total_deducted&limit=200')
sb_bal  = {p['display_name'].lower(): p for p in updated}

ws2 = wb['Corpus Summary']
rows2 = list(ws2.iter_rows(values_only=True))

ok, mismatch, unmatched_v = 0, 0, 0
print(f'  {"Player":<22} {"ExcelBal":>10} {"SBBal":>10} {"Diff":>8}  Note')
print('  ' + '-' * 70)

# Track which players played Apr 25
apr25_wid = 'W_2026_04_25'
apr25_players = set()
for rec in attendance:
    if rec['week_id'] == apr25_wid and rec['status'] == 'played':
        pid = rec['player_id']
        p = player_map.get(pid)
        if p and p['type'] in ('corpus', 'new'):
            apr25_players.add(pid)

for row in rows2[5:]:
    if not row[0] or not isinstance(row[0], (int, float)):
        continue
    name   = str(row[2]).strip() if row[2] else ''
    ex_bal = row[7]
    if not name or ex_bal is None:
        continue
    try:
        ex_bal = float(ex_bal)
    except (TypeError, ValueError):
        continue

    match = sb_bal.get(name.lower())
    if not match:
        for k, v in sb_bal.items():
            if name.lower() in k or k in name.lower():
                match = v
                break

    if not match:
        print(f'  {name:<22} {ex_bal:>10.2f}  {"NOT FOUND":>10}')
        unmatched_v += 1
        continue

    sb_b = float(match['corpus_balance'])
    diff = abs(sb_b - ex_bal)

    note = ''
    if match['id'] in apr25_players:
        fee = float(week_map.get(apr25_wid, {}).get('match_fee', 0) or 0)
        expected_diff = fee
        note = f'(played Apr25, -₹{fee:.0f})'

    flag = '  <-- MISMATCH' if diff > 5 and match['id'] not in apr25_players else ''
    print(f'  {name:<22} {ex_bal:>10.2f}  {sb_b:>10.2f}  {diff:>7.2f}{flag}  {note}')
    if diff > 5 and match['id'] not in apr25_players:
        mismatch += 1
    else:
        ok += 1

print()
print(f'  OK: {ok}  |  Mismatch (>5, excl Apr25): {mismatch}  |  Unmatched: {unmatched_v}')
